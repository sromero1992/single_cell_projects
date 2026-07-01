"""
pathway.py — TimeSCape pathway enrichment and activity scoring
==============================================================
Python equivalent of pathway_circadian.R.

Pipeline:
  1. pull_genesets()       — fetch MSigDB / custom gene sets via gseapy
  2. phase_bin_analysis()  — ORA per ZT acrophase bin (mirrors clusterProfiler::enricher)
  3. auc_score_cells()     — AUCell pathway activity per cell (via decoupler or fallback)
  4. pathway_cosinor()     — cosinor fit on per-cell pathway scores
  5. write_pathway_results() — Excel workbook (all + confident sheets)

Dependencies
------------
  pip install gseapy decoupler openpyxl

gseapy mirrors R's clusterProfiler; decoupler mirrors AUCell (Bioconductor).
Both are optional — graceful ImportError messages explain what to install.
"""

from __future__ import annotations

import warnings
import numpy as np
import pandas as pd
import scipy.sparse as sp
from typing import Optional

from .core  import estimate_phase_r
from .utils import bh_adjust


# ── 0. Fetch gene sets ────────────────────────────────────────────────────────

def pull_genesets(
    organism: str = "mouse",
    collection: str = "KEGG_2019_Mouse",
    min_size: int = 10,
    max_size: int = 500,
) -> dict[str, list[str]]:
    """
    Fetch gene sets from MSigDB or Enrichr via gseapy.

    Parameters
    ----------
    organism : str
        ``"mouse"`` or ``"human"`` (controls which Enrichr library to use).
    collection : str
        Name of the Enrichr gene set library. Examples:

        Mouse:
          - ``"KEGG_2019_Mouse"``
          - ``"Reactome_2022"`` (via gseapy Enrichr mouse libraries)
          - ``"GO_Biological_Process_2021"``

        Human:
          - ``"KEGG_2021_Human"``
          - ``"Reactome_2022"``
          - ``"GO_Biological_Process_2023"``

        Any library listed by ``gseapy.get_library_name()`` is accepted.
    min_size : int
        Minimum gene set size after symbol filtering (default 10).
    max_size : int
        Maximum gene set size (default 500).

    Returns
    -------
    dict mapping pathway name → list of gene symbols.

    Examples
    --------
    >>> gs = pull_genesets("mouse", "KEGG_2019_Mouse")
    >>> len(gs)   # e.g. 186
    """
    try:
        import gseapy as gp
    except ImportError:
        raise ImportError(
            "gseapy is required for pull_genesets(). "
            "Install it with:  pip install gseapy"
        )

    try:
        lib = gp.get_library(collection, organism=organism)
    except Exception as e:
        available = gp.get_library_name(organism=organism)
        raise ValueError(
            f"Could not fetch gene set library '{collection}' for organism '{organism}'.\n"
            f"Error: {e}\n"
            f"Available libraries: {available[:20]} ..."
        )

    # Filter by size
    filtered = {
        name: genes
        for name, genes in lib.items()
        if min_size <= len(genes) <= max_size
    }
    print(f"  pull_genesets: {len(filtered)} / {len(lib)} gene sets "
          f"({min_size}–{max_size} genes) from '{collection}'")
    return filtered


# ── 1. Phase-bin ORA ─────────────────────────────────────────────────────────

def phase_bin_analysis(
    conf_df: pd.DataFrame,
    genesets: dict[str, list[str]],
    universe: list[str],
    acrophase_col: str = "Acrophase_24",
    gene_col: str = "Genes",
    period: float = 24.0,
    bin_width: float = 2.0,
    n_top: int = 5,
    min_overlap: int = 3,
    min_bin_genes: int = 5,
    p_thresh: float = 0.05,
    use_padj: bool = True,
    exclude_patterns: list[str] | None = None,
) -> dict:
    """
    Bin confident circadian genes by acrophase, run ORA per ZT window.

    Mirrors R's ``phase_bin_analysis()``.  Genes from the same pathway that
    co-peak in the same ZT window are returned as phase-restricted gene sets
    (``phase_gs``) suitable for AUCell scoring.

    Parameters
    ----------
    conf_df : pd.DataFrame
        Confident circadian gene table (output of ``run_timescape``).
        Must contain ``gene_col`` and ``acrophase_col``.
    genesets : dict
        Named gene sets from ``pull_genesets()``.
    universe : list of str
        Background universe for ORA — use the list of all genes tested by
        ``run_timescape()``, not genome-wide, to avoid p-value inflation.
    period : float
        Circadian period in hours (default 24).
    bin_width : float
        ZT window width in hours (default 2).
    n_top : int
        Maximum top pathways to retain per bin (default 5).
    min_overlap : int
        Minimum gene overlap for a pathway hit (default 3).
    min_bin_genes : int
        Minimum genes in a bin to attempt ORA (default 5).
    p_thresh : float
        Significance threshold (default 0.05).
    use_padj : bool
        Use BH-adjusted p-value for filtering (default True).
    exclude_patterns : list of str, optional
        Pathway name substrings to exclude (case-insensitive).
        Default excludes DISEASE, VIRAL, INFECTION.

    Returns
    -------
    dict with keys:
        ``bin_table``   — conf_df with a ``phase_bin`` column added
        ``ora_results`` — dict of ORA hit DataFrames per active bin
        ``phase_gs``    — dict mapping "ZTxx-yy__PATHWAY" → gene list
    """
    try:
        import gseapy as gp
    except ImportError:
        raise ImportError(
            "gseapy is required for phase_bin_analysis(). "
            "Install it with:  pip install gseapy"
        )

    if exclude_patterns is None:
        exclude_patterns = ["DISEASE", "VIRAL", "INFECTION"]

    conf_df = conf_df.copy()
    acro_vals = conf_df[acrophase_col].values % period

    # Assign each gene to a ZT bin
    bins = np.arange(0, period, bin_width)
    bin_labels = [f"ZT{b:04.1f}-{b+bin_width:04.1f}" for b in bins]
    bin_idx = np.digitize(acro_vals, bins) - 1
    bin_idx = np.clip(bin_idx, 0, len(bins) - 1)
    conf_df["phase_bin"] = [bin_labels[i] for i in bin_idx]

    # Build background → set for fast lookup
    universe_set = set(universe)
    n_universe   = len(universe_set)

    ora_results: dict[str, pd.DataFrame] = {}
    phase_gs:   dict[str, list[str]]     = {}

    for b_idx, b_label in enumerate(bin_labels):
        bin_genes = conf_df.loc[conf_df["phase_bin"] == b_label, gene_col].tolist()
        if len(bin_genes) < min_bin_genes:
            continue

        bin_set = set(bin_genes)
        rows = []

        for pname, plist in genesets.items():
            # Filter by exclude patterns
            if any(pat.upper() in pname.upper() for pat in exclude_patterns):
                continue

            pset     = set(plist) & universe_set
            overlap  = bin_set & pset
            k = len(overlap)
            if k < min_overlap:
                continue

            K = len(bin_genes)       # genes in bin
            M = len(pset)            # pathway size in universe
            N = n_universe           # universe size

            # Hypergeometric p-value (Fisher exact upper tail)
            from scipy.stats import hypergeom
            pval = hypergeom.sf(k - 1, N, M, K)
            rows.append({
                "Bin":        b_label,
                "Pathway":    pname,
                "pvalue":     pval,
                "Overlap":    k,
                "PathwaySize": M,
                "BinGenes":   K,
                "Universe":   N,
                "RichFactor": k / M if M > 0 else 0.0,
                "Genes":      ";".join(sorted(overlap)),
            })

        if not rows:
            continue

        res_df = pd.DataFrame(rows)
        res_df["p_adj"] = bh_adjust(res_df["pvalue"].values)

        p_col = "p_adj" if use_padj else "pvalue"
        res_df = (res_df[res_df[p_col] < p_thresh]
                    .sort_values(p_col)
                    .head(n_top)
                    .reset_index(drop=True))

        if res_df.empty:
            continue

        ora_results[b_label] = res_df

        # Build phase-restricted gene sets for AUCell
        for _, row in res_df.iterrows():
            gs_key = f"{b_label}__{row['Pathway']}"
            # Genes: intersection of pathway genes AND bin genes
            pset = set(genesets.get(row["Pathway"], []))
            phase_gs[gs_key] = sorted(bin_set & pset)

    print(f"  phase_bin_analysis: {len(ora_results)} active bins, "
          f"{len(phase_gs)} phase-restricted gene sets")
    return {
        "bin_table":   conf_df,
        "ora_results": ora_results,
        "phase_gs":    phase_gs,
    }


# ── 2. AUCell scoring ─────────────────────────────────────────────────────────

def _aucell_numpy(
    X: np.ndarray,
    gene_names: list[str],
    genesets: dict[str, list[str]],
    auc_max_rank: float = 0.05,
) -> np.ndarray:
    """
    Minimal numpy AUCell implementation (fallback when decoupler is unavailable).

    X : genes × cells dense float array
    Returns: pathways × cells AUC score matrix.
    """
    n_genes, n_cells = X.shape
    auc_rank = max(1, int(round(auc_max_rank * n_genes)))

    gene_idx = {g: i for i, g in enumerate(gene_names)}

    # Pre-sort gene ranks per cell (argsort descending)
    order = np.argsort(-X, axis=0)  # genes × cells

    path_names = list(genesets.keys())
    scores = np.zeros((len(path_names), n_cells), dtype=np.float32)

    for pi, pname in enumerate(path_names):
        plist = genesets[pname]
        gs_idx = np.array([gene_idx[g] for g in plist if g in gene_idx], dtype=np.int32)
        if len(gs_idx) == 0:
            continue
        m = len(gs_idx)

        # For each cell: count how many pathway genes fall in the top auc_rank ranks
        # and compute AUC as the normalized sum of recoveries
        top_ranks = order[:auc_rank, :]   # auc_rank × n_cells

        for j in range(n_cells):
            top_set = set(top_ranks[:, j].tolist())
            # Ranks of pathway genes in this cell (position in sorted order)
            # AUC = sum over gs genes of (auc_rank - rank_in_top) / (auc_rank * m)
            hit_ranks = []
            for gi in gs_idx:
                # Find rank of this gene in the full ranking
                pos = np.where(order[:, j] == gi)[0]
                if len(pos) > 0 and pos[0] < auc_rank:
                    hit_ranks.append(pos[0])
            if hit_ranks:
                # AUC = area under recovery curve: sum(auc_rank - rank) / (auc_rank * m)
                scores[pi, j] = sum(auc_rank - r for r in hit_ranks) / (auc_rank * m)

    return scores


def auc_score_cells(
    adata,
    genesets: dict[str, list[str]],
    min_gs_size: int = 5,
    auc_max_rank: float = 0.05,
    layer: str | None = None,
    use_raw: bool = False,
) -> np.ndarray:
    """
    Compute AUCell pathway activity scores (pathways × cells).

    Uses ``decoupler.run_aucell()`` if available; falls back to a numpy
    implementation that is numerically equivalent for small datasets.

    Parameters
    ----------
    adata : AnnData
        Cells × genes. Use log-normalised counts (``adata.X`` after
        ``normalize_adata()`` or Scanpy preprocessing).
    genesets : dict
        Named gene sets from ``pull_genesets()`` or ``phase_bin_analysis()``.
    min_gs_size : int
        Minimum overlap between a gene set and the expression matrix to score
        that set (default 5).
    auc_max_rank : float
        Fraction of genes used as the AUC threshold (default 0.05 = top 5%).
    layer : str, optional
        AnnData layer to use (default: ``adata.X``).
    use_raw : bool
        Use ``adata.raw.X`` if True (default False).

    Returns
    -------
    auc_mat : np.ndarray  (pathways × cells)
        Row order matches ``list(genesets.keys())`` after size filtering.
        The row names can be recovered as a list from the returned matrix's
        attribute ``auc_mat.index`` (when decoupler is used, which returns
        a DataFrame; otherwise check the returned shape against filtered_keys).

    Notes
    -----
    The returned object is always a plain ``np.ndarray`` (pathways × cells)
    with a ``.pathway_names`` attribute listing the row labels.
    """
    import scipy.sparse as sp

    gene_names: list[str] = list(adata.var_names)

    # Get expression matrix
    if use_raw and adata.raw is not None:
        X_src = adata.raw.X
    elif layer is not None:
        X_src = adata.layers[layer]
    else:
        X_src = adata.X

    if sp.issparse(X_src):
        X_dense = X_src.toarray()
    else:
        X_dense = np.asarray(X_src)

    # X_dense is cells × genes — transpose to genes × cells for AUCell
    X_gc = X_dense.T.astype(np.float32)

    # Filter gene sets by overlap with expression matrix
    gene_set = set(gene_names)
    filtered_gs = {
        k: [g for g in v if g in gene_set]
        for k, v in genesets.items()
    }
    filtered_gs = {k: v for k, v in filtered_gs.items() if len(v) >= min_gs_size}

    print(f"  auc_score_cells: {len(filtered_gs)} / {len(genesets)} gene sets "
          f"with >={min_gs_size} genes in the expression matrix")
    if not filtered_gs:
        raise ValueError(
            "No gene sets overlap the expression matrix with min_gs_size genes. "
            "Check species / gene symbol format (mouse: Actb, human: ACTB)."
        )

    # Try decoupler first
    try:
        import decoupler as dc

        # Build long-format 'net' DataFrame required by decoupler
        rows = []
        for src, genes in filtered_gs.items():
            for g in genes:
                rows.append({"source": src, "target": g, "weight": 1.0})
        net = pd.DataFrame(rows)

        import anndata as ad
        adata_tmp = ad.AnnData(X=X_dense, var=pd.DataFrame(index=gene_names),
                               obs=pd.DataFrame(index=adata.obs_names))

        scores_df, _ = dc.run_aucell(
            adata_tmp, net,
            source="source", target="target", weight="weight",
            aucMaxRank=auc_max_rank,
            use_raw=False, verbose=False,
        )
        # scores_df: cells × pathways
        mat = scores_df.T.values.astype(np.float32)   # → pathways × cells
        path_order = list(scores_df.columns)

    except ImportError:
        warnings.warn(
            "decoupler not found — using built-in numpy AUCell. "
            "Install decoupler for faster, validated scoring:  pip install decoupler",
            ImportWarning,
            stacklevel=2,
        )
        mat = _aucell_numpy(X_gc, gene_names, filtered_gs, auc_max_rank=auc_max_rank)
        path_order = list(filtered_gs.keys())

    # Attach pathway names as an attribute for downstream use
    mat = np.asarray(mat, dtype=np.float32)
    # Store names as a plain list — callers can use this to build DataFrames
    result = mat
    result = np.ascontiguousarray(result)
    # Monkey-patch pathway names onto the array
    class _AnnotatedArray(np.ndarray):
        pass
    result_ann = result.view(_AnnotatedArray)
    result_ann.pathway_names = path_order

    return result_ann


# ── 3. Pathway-level cosinor ──────────────────────────────────────────────────

def pathway_cosinor(
    auc_mat: np.ndarray,
    pathway_names: list[str],
    obs: pd.DataFrame,
    tmeta: pd.DataFrame,
    target_ct: str,
    celltype_col: str = "cell_type",
    zt_col: str = "ZT_time",
    period: float = 24.0,
    test_type: str = "Ftest",
) -> dict:
    """
    Run cosinor on AUCell pathway activity scores for one cell type.

    Treats each pathway's per-cell AUC score exactly as ``run_timescape()``
    treats gene expression — the same ``estimate_phase_r()`` call, the same
    BH correction, the same confident threshold (raw p < 0.05 on both tests).

    Parameters
    ----------
    auc_mat : np.ndarray  (pathways × cells)
        From ``auc_score_cells()``.  Cell order must match ``obs``.
    pathway_names : list of str
        Row labels for ``auc_mat`` (e.g. ``auc_mat.pathway_names``).
    obs : pd.DataFrame
        Cell-level metadata (``adata.obs``).
    tmeta : pd.DataFrame
        ZT metadata with columns ``old_labels``, ``new_labels``, ``ZT_times``.
    target_ct : str
        Cell type to analyse (must match values in ``obs[celltype_col]``).
    celltype_col, zt_col : str
        Column names in ``obs``.
    period : float
        24.0 (default) or 12.0 for ultradian.
    test_type : str
        ``"Ftest"`` (default) or ``"LRT"``.

    Returns
    -------
    dict with keys:
        ``stats``    — pd.DataFrame, one row per pathway (same columns as T1)
        ``zt_means`` — pd.DataFrame, per-ZT mean AUC (pathways × ZT labels)
    """
    # ── Cell mask ─────────────────────────────────────────────────────────────
    ct_mask = obs[celltype_col].astype(str) == target_ct
    if ct_mask.sum() < 10:
        raise ValueError(
            f"Only {ct_mask.sum()} cells for '{target_ct}' — too few for pathway cosinor."
        )

    zt_strs = obs.loc[ct_mask, zt_col].astype(str).values

    # Map ZT strings → numeric hours
    old_to_num = dict(zip(tmeta["old_labels"].astype(str),
                          tmeta["ZT_times"].astype(float)))
    zt_num = np.array([old_to_num.get(z, np.nan) for z in zt_strs])

    valid_cells = np.isfinite(zt_num)
    zt_num   = zt_num[valid_cells]
    ct_idx   = np.where(ct_mask)[0][valid_cells]

    actual_times = np.sort(np.unique(zt_num))
    nzts = len(actual_times)
    if nzts < 4:
        raise ValueError(
            f"Only {nzts} ZT time points for '{target_ct}' — need at least 4."
        )

    auc_ct   = auc_mat[:, ct_idx]   # pathways × cells (this cell type)
    n_paths  = len(pathway_names)

    print(f"  pathway_cosinor: {n_paths} pathways × {len(ct_idx)} cells "
          f"({nzts} ZT points) for '{target_ct}'")

    records = []
    R0_rows = []

    for pi in range(n_paths):
        row_vals = auc_ct[pi, :]    # 1D: cells for this pathway
        Xg_zts = [row_vals[zt_num == z] for z in actual_times]

        fit = estimate_phase_r(Xg_zts, actual_times, period=period, test_type=test_type)

        acro24 = fit["acrophase"] % period if np.isfinite(fit["acrophase"]) else np.nan
        records.append({
            "Pathway":        pathway_names[pi],
            "Amp":            fit["amp"],
            "Abs_Amp":        abs(fit["amp"]) if np.isfinite(fit["amp"]) else np.nan,
            "Mesor":          fit["mesor"],
            "Acrophase":      fit["acrophase"],
            "Acrophase_24":   acro24,
            "Period":         fit["period"],
            "pvalue":         fit["p_value"],
            "Sine_corr":      fit["rho"],
            "pvalue_corr":    fit["p_value_macro"],
        })

        R0_rows.append(np.array([
            float(np.mean(Xg_zts[j])) if len(Xg_zts[j]) > 0 else np.nan
            for j in range(nzts)
        ]))

    stats_df = pd.DataFrame(records)

    # BH correction
    stats_df["pvalue_adj"]      = bh_adjust(stats_df["pvalue"].values)
    stats_df["pvalue_adj_corr"] = bh_adjust(stats_df["pvalue_corr"].values)

    # Sort: pvalue_adj_corr ↑, pvalue_adj ↑, Acrophase_24 ↑, Abs_Amp ↓
    stats_df = (stats_df
                .sort_values(["pvalue_adj_corr", "pvalue_adj",
                               "Acrophase_24", "Abs_Amp"],
                              ascending=[True, True, True, False])
                .reset_index(drop=True))

    # ZT means table
    zt_labels = [
        tmeta.loc[tmeta["ZT_times"] == t, "new_labels"].values
        for t in actual_times
    ]
    zt_labels = [
        lbl[0] if len(lbl) > 0 else f"ZT{int(t):02d}"
        for lbl, t in zip(zt_labels, actual_times)
    ]
    R0_mat = np.vstack(R0_rows)   # pathways × ZT
    zt_df  = pd.DataFrame(R0_mat, columns=zt_labels)
    zt_df.insert(0, "Pathway", [r["Pathway"] for r in records])

    conf_both = (stats_df["pvalue"] < 0.05) & (stats_df["pvalue_corr"] < 0.05)
    print(f"  Confident pathways (F+corr p<0.05): {conf_both.sum()}")

    return {"stats": stats_df, "zt_means": zt_df}


# ── 4. Write pathway results to Excel ────────────────────────────────────────

def write_pathway_results(
    results: dict,
    outpath: str,
    celltype: str = "",
) -> str:
    """
    Save pathway cosinor results to a formatted Excel workbook.

    Two sheets: All_Pathways (with confident rows highlighted) and
    Confident_Pathways.

    Parameters
    ----------
    results : dict
        Output of ``pathway_cosinor()``.
    outpath : str
        Full path for the ``.xlsx`` output file.
    celltype : str
        Label used in console messages.

    Returns
    -------
    outpath (str)
    """
    try:
        import openpyxl
        from openpyxl.styles import PatternFill, Font, Alignment
        from openpyxl.utils.dataframe import dataframe_to_rows
    except ImportError:
        raise ImportError(
            "openpyxl is required for write_pathway_results(). "
            "Install it with:  pip install openpyxl"
        )

    stats_df  = results["stats"]
    conf_mask = (stats_df["pvalue"] < 0.05) & (stats_df["pvalue_corr"] < 0.05)
    conf_df   = stats_df[conf_mask].reset_index(drop=True)

    wb = openpyxl.Workbook()

    # Styles
    hdr_fill  = PatternFill("solid", fgColor="2F4F8F")
    hdr_font  = Font(color="FFFFFF", bold=True)
    hdr_align = Alignment(horizontal="center")
    hi_fill   = PatternFill("solid", fgColor="FFF2CC")

    def _write_sheet(wb, name, df, highlight_rows=None):
        ws = wb.create_sheet(title=name)
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True)):
            ws.append(row)
            if r_idx == 0:
                for cell in ws[1]:
                    cell.fill   = hdr_fill
                    cell.font   = hdr_font
                    cell.alignment = hdr_align
            elif highlight_rows and r_idx in highlight_rows:
                for cell in ws[r_idx + 1]:
                    cell.fill = hi_fill
        # Auto-width (approximate)
        for col in ws.columns:
            max_len = max((len(str(cell.value or "")) for cell in col), default=8)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 40)

    # Sheet 1 — All pathways, confident rows highlighted
    conf_row_nums = set(np.where(conf_mask.values)[0].tolist())  # 0-indexed
    _write_sheet(wb, "All_Pathways", stats_df, highlight_rows=conf_row_nums)

    # Sheet 2 — Confident only
    _write_sheet(wb, "Confident_Pathways", conf_df)

    # Remove default empty sheet
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    wb.save(outpath)
    print(f"  Pathway results saved → {outpath}")
    return outpath
